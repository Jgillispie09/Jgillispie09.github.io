import requests
from bs4 import BeautifulSoup
import pandas as pd
import datetime
from openpyxl import load_workbook

dt = datetime.datetime.now()
dt = str(dt).replace(' ', '__')
dt = dt.replace(':', '_')

#The purpose of this webcrawler is to act as an analytical tool for video game data (name, publisher, release date, review score, player count[if available])
url1 = "https://www.ign.com/reviews/games"
url2 = "https://www.gamespot.com/games/reviews/"
url3 = "https://www.metacritic.com/browse/game/all/all/current-year/metascore/?platform=ps5&platform=xbox-series-x&platform=nintendo-switch&platform=pc&platform=mobile&page=1"
url4 = "https://store.steampowered.com/search/?sort_by=Reviews_DESC&supportedlang=english&filter=popularnew&ndl=1"
#Creates a header to help the server understand which user-agent is requesting data, this helps the crawler not get blocked and misinterpreted as another type of bot
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

response1 = requests.get(url1, headers=headers)
response2 = requests.get(url2, headers=headers)
response3 = requests.get(url3, headers=headers)
response4 = requests.get(url4, headers=headers)

response_counter = 0

while response_counter != 4:
    if response1.status_code == 200:
        print("Successful server response from IGN.")
        response_counter += 1
    else:
        print("An error has occurred with IGN!")

    if response2.status_code == 200:
        print("Successful server response from GameSpot.")
        response_counter += 1
    else:
        print("An error has occurred with GameSpot!")

    if response3.status_code == 200:
        print("Successful server response from MetaCritic.")
        response_counter += 1
    else:
        print("An error has occurred with MetaCritic!")

    if response4.status_code == 200:
        print("Successful server response from Steam.")
        response_counter += 1
    else:
        print("An error has occurred!")

    if response_counter != 4:
        print("Error occured with server response, exiting program.")
        exit()

html_content1 = response1.content
html_content2 = response2.content
html_content3 = response3.content
html_content4 = response4.content

soup1 = BeautifulSoup(html_content1, 'html.parser')
soup2 = BeautifulSoup(html_content2, 'html.parser')
soup3 = BeautifulSoup(html_content3, 'html.parser')
soup4 = BeautifulSoup(html_content4, 'html.parser')

game_rows1 = soup1.find_all('div', class_="content-item jsx-2981448069 row divider")
game_rows2 = soup2.find_all('div', class_='card-item base-flexbox flexbox-align-center width-100 border-bottom-grayscale--thin')
game_rows3 = soup3.find_all('div', class_='c-finderProductCard c-finderProductCard-game')
game_rows4 = soup4.find_all('div', class_='responsive_search_name_combined')

game_data1 = []
game_data2 = []
game_data3 = []
game_data4 = []

for game in game_rows1:
    title = game.find('a', class_='caption jsx-1762799490 data small').text.strip() if game.find('a', class_='caption jsx-1762799490 data small') else "N/A"
    rating = game.find('span', class_='hexagon-content-wrapper').text.strip() if game.find('span', class_='hexagon-content-wrapper') else "N/A"
    excerpt = game.find('div', class_='interface jsx-326752785 jsx-957202555 item-subtitle small').text.strip() if game.find('div', class_='interface jsx-326752785 jsx-957202555 item-subtitle small') else "N/A"

    excerpt = excerpt.split('- ')
    excerpt = excerpt[1]

    game_data1.append({
        "Title": title,
        "Review Score": rating,
        "Excerpt" : excerpt
    })

for game in game_rows2:
    title = game.find('h4', class_='card-item__title').text.strip() if game.find('h4', class_='card-item__title') else "N/A"
    rating = game.find('div', class_='review-ring-score__score text-bold').text.strip() if game.find('div', class_='review-ring-score__score text-bold') else "N/A"
    review = game.find('span', class_='review-ring-score__text review-ring-score__text--xsmall text-center').text.strip() if game.find('span', class_='review-ring-score__text review-ring-score__text--xsmall text-center') else "N/A"
    
    game_data2.append({
        "Article Title": title,
        "Review Score": rating,
        "Quality": review
    })

for game in game_rows3:
    title = game.find('h3', class_='c-finderProductCard_titleHeading').text.strip() if game.find('h3', class_='c-finderProductCard_titleHeading') else 'N/A'
    publish_date = game.find('span', class_='u-text-uppercase').text.strip() if game.find('span', class_='u-text-uppercase') else "N/A"
    rating = game.find('div', class_='c-siteReviewScore u-flexbox-column u-flexbox-alignCenter u-flexbox-justifyCenter g-text-bold c-siteReviewScore_green g-color-gray90 c-siteReviewScore_xsmall').text.strip() if game.find('div', class_='c-siteReviewScore u-flexbox-column u-flexbox-alignCenter u-flexbox-justifyCenter g-text-bold c-siteReviewScore_green g-color-gray90 c-siteReviewScore_xsmall') else "N/A"

    game_data3.append({
        "Title": title,
        "Publish Date": publish_date,
        "MetaScore": rating
    })

for game in game_rows4:
    title = game.find('span', class_='title').text.strip() if game.find('span', class_='title') else 'N/A'
    release_date = game.find('div', class_='col search_released responsive_secondrow').text.strip() if game.find('div', class_='col search_released responsive_secondrow') else "N/A"
    review_summary = game.find('span', class_='search_review_summary positive')
    if review_summary and review_summary.has_attr('data-tooltip-html'):
        review = review_summary['data-tooltip-html'].split('<br>')[0].strip()
        review_score = review_summary['data-tooltip-html'].split('<br>')[1].strip()
        review_score = review_score.split(' ')[0].strip()
    else:
        review = "N/A"
    
    game_data4.append({
        "Title": title,
        "Release Date": release_date,
        "Review Summary": review,
        "Review Score": review_score
    })

for game in game_data1:
    print(game)

for game in game_data2:
    print(game)

for game in game_data3:
    print(game)

for game in game_data4:
    print(game)

df1 = pd.DataFrame(game_data1)
df2 = pd.DataFrame(game_data2)
df3 = pd.DataFrame(game_data3)
df4 = pd.DataFrame(game_data4)

output_file = "Game Data Compiled " + dt + ".xlsx"

csv1 = "IGN Review Data " + dt + ".xlsx"
csv2 = "GameSpot Review Data " + dt + ".xlsx"
csv3 = "MetaCritic Review Data " + dt + ".xlsx"
csv4 = "Steam Review Data " + dt + ".xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:    
    df1.to_excel(writer, sheet_name="IGN Review Data", index=False)
    df2.to_excel(writer, sheet_name="GameSpot Review Data", index=False)
    df3.to_excel(writer, sheet_name="MetaCritic Review Data", index=False)
    df4.to_excel(writer, sheet_name="Steam Review Data", index=False)

workbook = load_workbook(output_file)

for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        worksheet.column_dimensions[column_letter].width = max_length + 2

workbook.save(output_file)

print(f"Data Extracted to Excel file as {output_file}.")